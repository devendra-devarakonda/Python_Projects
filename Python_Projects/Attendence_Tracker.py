import openpyxl
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# Load the Excel sheet
book = openpyxl.load_workbook(r'D:\HHO-Volunteers\Book1.xlsx')
sheet = book['Sheet1']

# Counting the number of rows (students) and columns (subjects)
r = sheet.max_row
c = sheet.max_column


m1 = "I hope you are doing well. I wanted to bring to your attention that your attendance for the 'IOR' class has fallen below the minimum required 80%.\n As per the course policy, students are required to maintain at least '80%' attendance to be eligible for examinations.\n I kindly request that you make an effort to attend all future classes in order to improve your attendance percentage.Please consider this as a formal warning. If your attendance does not improve, further action may be taken as per the institution's attendance policies.\n If you have any questions or concerns, feel free to reach out to me."
m2 = "I hope you are doing well. I wanted to bring to your attention that your attendance for the 'COA' class has fallen below the minimum required 80%.\n As per the course policy, students are required to maintain at least '80%' attendance to be eligible for examinations.\n I kindly request that you make an effort to attend all future classes in order to improve your attendance percentage.Please consider this as a formal warning. If your attendance does not improve, further action may be taken as per the institution's attendance policies.\n If you have any questions or concerns, feel free to reach out to me."
m3 = "I hope you are doing well. I wanted to bring to your attention that your attendance for the 'DSP' class has fallen below the minimum required 80%.\n As per the course policy, students are required to maintain at least '80%' attendance to be eligible for examinations.\n I kindly request that you make an effort to attend all future classes in order to improve your attendance percentage.Please consider this as a formal warning. If your attendance does not improve, further action may be taken as per the institution's attendance policies.\n If you have any questions or concerns, feel free to reach out to me."
m4 = "I hope you are doing well. I wanted to bring to your attention that your attendance for the 'WT' class has fallen below the minimum required 80%.\n As per the course policy, students are required to maintain at least '80%' attendance to be eligible for examinations.\n I kindly request that you make an effort to attend all future classes in order to improve your attendance percentage.Please consider this as a formal warning. If your attendance does not improve, further action may be taken as per the institution's attendance policies.\n If you have any questions or concerns, feel free to reach out to me."
m5 = "I hope you are doing well. I wanted to bring to your attention that your attendance for the 'CD' class has fallen below the minimum required 80%.\n As per the course policy, students are required to maintain at least '80%' attendance to be eligible for examinations.\n I kindly request that you make an effort to attend all future classes in order to improve your attendance percentage.Please consider this as a formal warning. If your attendance does not improve, further action may be taken as per the institution's attendance policies.\n If you have any questions or concerns, feel free to reach out to me."



# Dictionary to hold total number of classes for each subject
total_classes = {
    1: 30,
    2: 30,
    3: 30,
    4: 30,
    5: 30,
}

def savefile():
    book.save(r'D:\HHO-Volunteers\Book1.xlsx')
    print("Saved!")

def calculate_attendance(student, subject):
    if subject == 1:
        total_days = total_classes[1]
        absent_days = sheet.cell(row=student, column=3).value
    elif subject == 2:
        total_days = total_classes[2]
        absent_days = sheet.cell(row=student, column=4).value
    elif subject == 3:
        total_days = total_classes[3]
        absent_days = sheet.cell(row=student, column=5).value
    elif subject == 4:
        total_days = total_classes[4]
        absent_days = sheet.cell(row=student, column=6).value
    else:
        total_days = total_classes[5]
        absent_days = sheet.cell(row=student, column=7).value
    
    present_days = total_days - absent_days
    percentage = (present_days / total_days) * 100
    return percentage

def send_warning_email(student_email, message):
    from_id = 'forbingtwo@gmail.com'
    pwd = 'kiqt jupc bzbg octn'
    s = smtplib.SMTP('smtp.gmail.com', 587, timeout=120)
    s.starttls()
    s.login(from_id, pwd)

    msg = MIMEMultipart()
    msg['From'] = from_id
    msg['To'] = student_email
    msg['Subject'] = 'Attendance Warning'

    msg.attach(MIMEText(message, 'plain'))

    content = msg.as_string() # Convert the message to a string format
    s.sendmail(from_id, student_email, content)
    s.quit()

    print(f"Mail sent to {student_email}")

def check_attendance(no_of_days, row_num, b):
    global staff_mails
    for student in range(0, len(row_num)):
        percentage = calculate_attendance(row_num[student], b)
        
        if percentage < 80:
            student_email = sheet.cell(row=row_num[student], column=2).value
            if b == 1:
                send_warning_email(student_email, m1)
            elif b == 2:
                send_warning_email(student_email, m2)
            elif b == 3:
                send_warning_email(student_email, m3)
            elif b == 4:
                send_warning_email(student_email, m4)
            else:
                send_warning_email(student_email, m5)

abc=int(input("Having Any Absentees Today : \n (1)-->Yes \n (2)-->No"))
while abc == 1:
    print("1--->IOR\n2--->COA\n3--->DSP\n4-->WT\n-->CD")
    y = int(input("Enter subject: "))

    no_of_absentees = int(input('Number of absentees: '))
    x = list(map(int, (input('Roll numbers: ').split(' ')))) if no_of_absentees > 1  else [int(input('Roll no: '))]
   
        

    row_num = []
    no_of_days = []

    for student in x:
        for i in range(2, r+1):
            if y == 1:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=3).value
                    m += 1
                    sheet.cell(row=i, column=3).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)
                    total_classes[1] += 1

            elif y == 2:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=4).value
                    m += 1
                    sheet.cell(row=i, column=4).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)
                    total_classes[2] += 1

            elif y == 3:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=5).value
                    m += 1
                    sheet.cell(row=i, column=5).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)
                    total_classes[3] += 1


            elif y == 4:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=6).value
                    m += 1
                    sheet.cell(row=i, column=6).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)
                    total_classes[4] += 1


            elif y == 5:
                if sheet.cell(row=i, column=1).value == student:
                    m = sheet.cell(row=i, column=7).value
                    m += 1
                    sheet.cell(row=i, column=7).value = m
                    savefile()
                    no_of_days.append(m)
                    row_num.append(i)
                    total_classes[5] += 1

    check_attendance(no_of_days, row_num, y)
    
    resp = int(input('Another subject? 1--->Yes, 0--->No: '))
    if resp == 0:
        break
